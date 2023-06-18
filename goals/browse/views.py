from re import M
from django.shortcuts import render
from .models import Goal, Quarter, Summary
from .forms import AddGoalForm, SummaryForm, EditSummaryForm
from django.contrib.auth.models import User
import datetime
from django.http import JsonResponse, HttpResponse
from django.forms.models import model_to_dict
from .validators import goal_validator
from django.contrib.auth.decorators import login_required
from datetime import date
from django.utils.timezone import localtime
from django.db.models import Q
from django.contrib.auth.models import Permission
from .database_client import true_converter, edit_goal, \
    send_message, edit_summary
from django.contrib.auth.decorators import user_passes_test
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from tempfile import NamedTemporaryFile
from urllib.parse import quote
from .models import CHOICES_BLOCK
import openpyxl.utils.cell
from users.views import get_notifications_once, get_user_name
from users.models import Image


@login_required(login_url='/user/login/')
def browse(request):
    return render(request, 'browse/browse.html')


@login_required(login_url='/user/login/')
def browse_add(request):
    return render(request, 'browse/add.html')


@login_required(login_url='/user/login/')
def approve_goal(request):
    if request.user.has_perm('browse.change_goal'):
        return render(request, 'browse/approve.html')


@user_passes_test(lambda u: u.is_superuser, login_url='/user/login/')
@login_required(login_url='/user/login/')
def summary(request):
    return render(request, 'browse/summary.html')


@user_passes_test(lambda u: u.is_superuser, login_url='/user/login/')
@login_required(login_url='/user/login/')
def browse_summary(request):
    return render(request, 'browse/browse_summary.html')


@login_required(login_url='/user/login/')
def editing(request):
    if request.method == "POST":
        goal = Goal.objects.get(id=request.POST.get('goal_id'))
        if (request.user == goal.owner_id or
                request.user.is_superuser or
                len(request.user.groups.all() &
                    goal.owner_id.groups.all()) > 0 and
                request.user.has_perm('browse.change_goal')):
            if not goal_validator(request):
                return JsonResponse({'status': 'validation error'})
            edit_goal(request, goal)
            return JsonResponse({'status': 'ok'})
        else:
            return JsonResponse({'status': '403 forbidden'})


@login_required(login_url='/user/login/')
def chatting(request):
    if request.method == "POST":
        goal = Goal.objects.get(id=request.POST.get('goal_id'))
        send_message(request, goal)
        return JsonResponse({'status': 'ok'})


@login_required(login_url='/user/login/')
def get_chat(request):
    goal = Goal.objects.get(id=request.GET.get('goal_id'))
    chats = goal.chat_set.all()
    chat_dict = {'chat': []}
    for chat in chats:
        chat_dict['chat'].append({'text': chat.message,
                                  'time': localtime(chat.created_at),
                                  'name': chat.owner_id.get_full_name(),
                                  'user_id': chat.owner_id.id})
    return JsonResponse(chat_dict)


@login_required(login_url='/user/login/')
def get_goal(request):
    goal = Goal.objects.get(id=request.GET.get('goal_id'))
    goal_dict = model_to_dict(goal)
    chats = goal.chat_set.all()
    histories = goal.history_set.all()
    goal_dict['chat'] = []
    goal_dict['history'] = []
    for chat in chats:
        goal_dict['chat'].append({'text': chat.message,
                                  'user_id': chat.owner_id.id,
                                  'time': localtime(chat.created_at),
                                  'name': chat.owner_id.get_full_name()})
    for hist in histories:
        hist_fc = []
        for fc in hist.fieldchange_set.all():
            hist_fc.append(model_to_dict(fc))
        goal_dict['history'].append({'name': hist.owner_id.get_full_name(),
                                     'owner_id': hist.owner_id.id,
                                     'time': localtime(hist.created_at),
                                     'field_changes': hist_fc})
    goal_dict['user_name'] = User.objects.get(
        id=goal_dict['owner_id']).get_full_name()
    goal_dict['admin_rights'] = \
        len(request.user.groups.all() &
            goal.owner_id.groups.all()) > 0 and \
        request.user != goal.owner_id and \
        request.user.has_perm('browse.change_goal')
    if request.user.is_superuser:
        goal_dict['admin_rights'] = True
    goal_dict['rights'] = goal.owner_id == request.user
    if goal_dict['admin_rights']:
        goal_dict['rights'] = True
    return JsonResponse(goal_dict)


@login_required(login_url='/user/login/')
def get_goals_by_filter(request):
    goals = Goal.objects.all()
    filters = {
        'block': request.GET.get('block', None),
        'sorting': request.GET.get('sort', None),
        'planned': request.GET.get('planned', None),
        'done': request.GET.get('done', None),
        'picked': request.GET.get('picked', None),
        'my': true_converter.get(request.GET.get('self'), None),
        'current': true_converter.get(request.GET.get('current'), None),
        'approve': true_converter.get(request.GET.get('approve'), None),
        'search': request.GET.get('search', None),
        'quarters': request.GET.getlist('quarter[]'),
        'summary_id': request.GET.get('summary_id', None),
        'owner_id': request.GET.get('owner_id', None),
        'sorting_group': request.GET.get('sort_group')
    }

    if filters['block'] == 'Все':
        filters['block'] = None
    if filters['sorting'] == 'Все':
        filters['sorting'] = None
    if filters['planned'] == 'Все':
        filters['planned'] = None
    if filters['done'] == 'Все':
        filters['done'] = None
    if filters['picked'] == 'Все':
        filters['picked'] = None

    if filters['summary_id']:
        summary = Summary.objects.get(id=filters['summary_id'])
        all_goals = Goal.objects.filter(quarter=summary.quarter,
                                        block=summary.block)
        intersection = summary.goals.all()
        picked_filtered_goals = all_goals.exclude(pk__in=intersection)
        goals = picked_filtered_goals

    if filters['approve']:
        if request.user.is_superuser:
            perm = Permission.objects.get(codename='change_goal')
            goals = goals.filter(Q(owner_id__groups__permissions=perm) |
                                 Q(owner_id__user_permissions=perm)).distinct()
        else:
            goals = goals.filter(owner_id__groups__in=request.user.groups.all())
            goals = goals.exclude(owner_id=request.user)

    if filters['block']:
        goals = goals.filter(block=filters['block'])

    if filters['sorting']:
        if filters['sorting'] == 'weight':
            goals = goals.order_by('-'+filters['sorting'])
        else:
            goals = goals.order_by(filters['sorting'])

    if filters['planned']:
        goals = goals.filter(planned=True
                             if filters['planned'] == 'Запланированные'
                             else False)
    if filters['my']:
        goals = goals.filter(owner_id=request.user)

    if filters['search']:
        search = filters['search'].strip()
        search_filters = Q(name__icontains=search) | Q(description__icontains=search) | Q(current_result__icontains=search)
        splitted_search = search.split()
        if len(splitted_search) == 1:
            search_filters |= Q(owner_id__first_name__icontains=search) | Q(owner_id__last_name__icontains=search)
        elif len(splitted_search) == 2:
            first_name, last_name = splitted_search[0], splitted_search[1]
            search_filters |= Q(owner_id__first_name__icontains=first_name, owner_id__last_name__icontains=last_name) | Q(owner_id__first_name__icontains=last_name, owner_id__last_name__icontains=first_name)

        goals = goals.filter(search_filters)
    if filters['quarters']:
        goals = goals.filter(quarter__in=filters['quarters'])

    if filters['current'] in [True, False]:
        goals = goals.filter(current=filters['current'])
    if filters['done']:
        goals = goals.filter(isdone=True if filters['done'] == 'Выполненные' else False)
    if filters['picked']:
        if filters['picked'] == 'Включено':
            goals = goals.exclude(summaries_count=0)
        else:
            goals = goals.filter(summaries_count=0)
    if filters['owner_id']:
        goals = goals.filter(owner_id=filters['owner_id'])

    users_dict = {}
    users_list = []
    data = list(goals.values('name', 'weight', 'isdone', 'owner_id', 'block', 'id', 'quarter', 'summaries_count'))
    for item in data:
        user = User.objects.get(id=item['owner_id'])
        if user.id not in users_dict:
            users_dict[user.id] = 1
        else:
            users_dict[user.id] += 1
        item['owner'] = user.get_full_name()
        item['owner_id'] = user.id
        item['picked'] = item['summaries_count'] > 0
    for i in users_dict:
        user = User.objects.get(id=i)
        users_list.append({'owner_id': user.id, 'name': user.get_full_name(),
                           'count': users_dict[i]})
    if filters['sorting_group']:
        if filters['sorting_group'] == 'count':
            users_list.sort(key=lambda x: x['count'], reverse=True)
        else:
            users_list.sort(key=lambda x: x['name'], reverse=True)
    answer_dict = {'goals': data, 'groups': users_list}

    return JsonResponse(answer_dict, safe=False)


@login_required(login_url='/user/login/')
def add_goal(request):
    if request.method == 'POST':
        form = AddGoalForm(request.POST)
        if form.is_valid():
            goal = Goal(owner_id=request.user,
                        name=request.POST.get('name'),
                        description=request.POST.get('description'),
                        block=request.POST.get('block'),
                        quarter=request.POST.get('quarter'),
                        weight=float(request.POST.get('weight')),
                        current=False,
                        current_result='',
                        planned=true_converter[request.POST.get('planned')],
                        mark=0,
                        fact_mark=0,
                        isdone=False,
                        )
            goal.save()
            return JsonResponse({'status': 'ok'})
        else:
            return JsonResponse({'status': 'validation error'})


@login_required(login_url='/user/login/')
def delete_goal(request) -> JsonResponse | None:
    if request.method == 'POST':
        goal = Goal.objects.get(id=request.POST.get('goal_id'))
        if (request.user == goal.owner_id or
                request.user.is_superuser or
                len(request.user.groups.all() &
                    goal.owner_id.groups.all()) > 0 and
                request.user.has_perm('browse.change_goal')):
            if not goal.current:
                goal.delete()
                return JsonResponse({'status': 'ok'})
            else:
                return JsonResponse({'status':
                                    'Утверждённые задачи нельзя удалить'})
        else:
            return JsonResponse({'status': '403 forbidden'})


@user_passes_test(lambda u: u.is_superuser, login_url='/user/login/')
@login_required(login_url='/user/login/')
def get_summaries(request) -> JsonResponse:
    block = request.GET.get('block') \
        if request.GET.get('block') != 'Все' else None
    quarter = request.GET.get('quarter')
    search = request.GET.get('search')
    summaries = Summary.objects.all()
    if block:
        summaries = summaries.filter(block=block)
    if quarter:
        summaries = summaries.filter(quarter=quarter)
    if search:
        search = search.strip()
        q1 = summaries.filter(name__icontains=search)
        q2 = summaries.filter(plan__icontains=search)
        q3 = summaries.filter(fact__icontains=search)
        summaries = q1 | q2 | q3
    return JsonResponse(list(summaries.values()), safe=False)


@user_passes_test(lambda u: u.is_superuser, login_url='/user/login/')
@login_required(login_url='/user/login/')
def get_summary(request) -> JsonResponse:
    summary = Summary.objects.get(id=request.GET.get('summary_id'))
    summary_dict = model_to_dict(summary)
    goal_ids = []
    for goal in summary_dict['goals']:
        g = model_to_dict(goal)
        g['owner'] = goal.owner_id.get_full_name()
        goal_ids.append(g)
    summary_dict['goals'] = goal_ids
    return JsonResponse(summary_dict)


@user_passes_test(lambda u: u.is_superuser, login_url='/user/login/')
@login_required(login_url='/user/login/')
def add_summary(request) -> JsonResponse | None:
    if request.method == 'POST':
        form = SummaryForm(request.POST)
        if form.is_valid():
            summary = Summary(plan=request.POST.get('plan'),
                              fact=request.POST.get('fact'),
                              block=request.POST.get('block'),
                              quarter=request.POST.get('quarter'),
                              name=request.POST.get('name'),
                              average_mark=0)
            summary.save()
            summary.goals.set(Goal.objects.filter(
                pk__in=request.POST.getlist('goals[]')))
            summary.save()
            goals = request.POST.getlist('goals[]')
            for id in goals:
                goal = Goal.objects.get(id=id)
                goal.summaries_count += 1
                goal.save()
            return JsonResponse({'status': 'ok'})
        else:
            return JsonResponse({'status': 'validation error',
                                'error': form.errors})


@user_passes_test(lambda u: u.is_superuser, login_url='/user/login/')
@login_required(login_url='/user/login/')
def editing_summary(request) -> JsonResponse | None:
    if request.method == "POST":
        summary = Summary.objects.get(id=request.POST.get('summary_id'))
        if request.user.is_superuser:
            form = EditSummaryForm(request.POST)
            if form.is_valid():
                edit_summary(request, summary)
                return JsonResponse({'status': 'ok'})
            else:
                return JsonResponse({'status': 'validation error',
                                     'error': form.errors})
        else:
            return JsonResponse({'status': '403 forbidden'})


@user_passes_test(lambda u: u.is_superuser, login_url='/user/login/')
@login_required(login_url='/user/login/')
def delete_summary(request) -> JsonResponse | None:
    if request.method == 'POST':
        summary = Summary.objects.get(id=request.POST.get('summary_id'))
        summary.delete()
        return JsonResponse({'status': 'ok'})


@user_passes_test(lambda u: u.is_superuser, login_url='/user/login/')
@login_required(login_url='/user/login/')
def download_summaries(request) -> HttpResponse:
    wb = openpyxl.Workbook()
    quarter = request.GET.get('quarter')
    ws = wb.active
    ws.title = quarter
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    green_fill = PatternFill(start_color="C6EFCE",
                             end_color="C6EFCE",
                             fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C",
                              end_color="FFEB9C",
                              fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE",
                           end_color="FFC7CE",
                           fill_type="solid")
    top_left_alignment = Alignment(wrapText=True,
                                   vertical='top',
                                   horizontal='left')
    ws['A1'] = 'Легенда'
    ws['A1'].alignment = top_left_alignment

    ws['A2'] = 'Задача выполнена с ожидаемым (и выше) результатом'
    ws['A2'].fill = green_fill
    ws['A2'].alignment = top_left_alignment
    ws.row_dimensions[2].height = 60

    ws['A3'] = 'Задача выполнена, но не полностью'
    ws['A3'].fill = yellow_fill
    ws['A3'].alignment = top_left_alignment
    ws.row_dimensions[3].height = 60

    ws['A4'] = 'Задача не выполнена'
    ws['A4'].fill = red_fill
    ws['A4'].alignment = top_left_alignment
    ws.row_dimensions[4].height = 60

    ws['C5'] = f'Сводка за {quarter}'.upper()
    ws['C5'].font = Font(bold=True, size=26)
    ws['C5'].alignment = Alignment(horizontal='center')
    ws.merge_cells('C5:G5')
    ws.row_dimensions[5].height = 60
    max_block_len = 0
    for i, block in enumerate(CHOICES_BLOCK[1:]):
        letter = openpyxl.utils.cell.get_column_letter(i+1)
        cell_title = f'{letter}6'
        ws[cell_title] = block[0]
        ws[cell_title].border = thin_border
        ws[cell_title].font = Font(bold=True)
        ws[cell_title].alignment = Alignment(horizontal='center')
        summaries = Summary.objects.filter(block=block[0], quarter=quarter)
        for j, summary in enumerate(summaries):
            cell = f'{letter}{j+7}'
            ws[cell] = \
                f'{summary.name}\nПЛАН: {summary.plan}\nФАКТ: {summary.fact}'
            ws[cell].alignment = top_left_alignment
            ws[cell].border = thin_border
            if summary.average_mark >= 50:
                ws[cell].fill = yellow_fill
            if summary.average_mark >= 70:
                ws[cell].fill = green_fill
            if summary.average_mark < 50:
                ws[cell].fill = red_fill
        max_block_len = max(max_block_len, len(summaries))
        ws.column_dimensions[letter].width = 35

    rows = ws[f'A6:I{max_block_len+6}']
    for row in rows:
        for cell in row:
            cell.border = thin_border

    ws.auto_filter.ref = f'A6:I{max_block_len+6}'

    with NamedTemporaryFile(delete=True) as tmp_file:
        wb.save(tmp_file.name)
        title = f'Сводка за {quarter}'
        with open(tmp_file.name, 'rb') as f:
            response = HttpResponse(f.read(),
                                    content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = \
                f'attachment; filename*=UTF-8\'\'{quote(title)}.xlsx'
            return response


@login_required(login_url='/user/login/')
def start_init(request) -> JsonResponse:
    d = get_quarters(request)
    notify = get_notifications_once(request)
    username = get_user_name(request)
    images = Image.objects.all()
    users = User.objects.all()
    d['images'] = {}
    for user in users:
        d['images'][user.id] = 'user_logo.png'
    for img in images:
        try:
            d['images'][img.user.id] = img.image.url.split('/')[::-1][0]
        except:
            pass
    d.update(notify)
    d.update(username)
    return JsonResponse(d)


def get_quarters(request):
    now = datetime.datetime.now()
    current_year = date.today().year
    quarter_of_the_year = int(f'{(now.month-1)//3+1}')
    current_quarter_string = f'{quarter_of_the_year} квартал {current_year}'
    if len(Quarter.objects.all()) == 0:
        new_quarter = Quarter(quarter=current_quarter_string)
        new_quarter.save()
    last_quarter_model = Quarter.objects.last().quarter.split()
    last_quarter = int(last_quarter_model[0])
    last_year = int(last_quarter_model[2])

    if current_quarter_string == ' '.join(last_quarter_model):
        last_quarter += 1
        if last_quarter == 5:
            last_quarter = 1
            last_year += 1
        new_quarter = Quarter(quarter=f'{last_quarter} квартал {last_year}')
        new_quarter.save()

    all_quarters = [(quarter.quarter, quarter.quarter)
                    for quarter in Quarter.objects.all()]
    choices = sorted([quarter[0] for quarter in all_quarters],
                     key=lambda quarter: (quarter.split()[2],
                                          quarter.split()[0]), reverse=True)

    data = {'quarters': choices, 'current_quarter': current_quarter_string}

    return data
