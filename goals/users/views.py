from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from .models import Notification, Image
from browse.models import Goal
from django.utils.timezone import localtime
import openpyxl
from tempfile import NamedTemporaryFile
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from urllib.parse import quote
import base64
from django.core.files.base import ContentFile
from .validators import validate_image_extension, file_size
from django.core.exceptions import ValidationError
from django.contrib.auth.models import User
import random
from string import ascii_lowercase, ascii_uppercase


def get_notifications_once(request):
    active_notifi = []

    for notifi in Notification.objects.all():
        if notifi.is_active():
            active_notifi.append(notifi.id)

    notifi = Notification.objects.filter(id__in=active_notifi)
    notifi = notifi.filter(user=request.user).order_by('-created_at')

    notifi_list = list(notifi.values())
    for notifi in notifi_list:
        notifi['goal_name'] = Goal.objects.get(id=notifi['goal_id']).name
        notifi['created_at'] = localtime(notifi['created_at'])

    return {'notify': notifi_list}


@login_required(login_url='/user/login/')
def get_notifications(request):
    active_notifi = []

    for notifi in Notification.objects.all():
        if notifi.is_active():
            active_notifi.append(notifi.id)

    notifi = Notification.objects.filter(id__in=active_notifi)
    notifi = notifi.filter(user=request.user).order_by('-created_at')

    notifi_list = list(notifi.values())
    for notifi in notifi_list:
        notifi['goal_name'] = Goal.objects.get(id=notifi['goal_id']).name
        notifi['created_at'] = localtime(notifi['created_at'])

    return JsonResponse(notifi_list, safe=False)


@login_required(login_url='/user/login/')
def read_notification(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        if id == 'all':
            notifi = Notification.objects.filter(user=request.user)
            for n in notifi:
                n.is_read = True
        else:
            notifi = Notification.objects.get(id=id)
            notifi.is_read = True
        notifi.save()
        return JsonResponse({'status': 'ok'})


@login_required(login_url='/user/login/')
def upload_image(request):
    if request.method == 'POST':
        file = request.POST.get('file')
        format, imgstr = request.POST.get('file').split(';base64,')
        ext = format.split('/')[-1]
        random_string = ''.join(random.choice(ascii_uppercase+ascii_lowercase) for _ in range(20))
        data = ContentFile(base64.b64decode(imgstr),
                           name=random_string + '.' + ext)
        try:
            validate_image_extension(ext)
            file_size(file)
        except (ValidationError):
            return JsonResponse({'status': 'validation error'})
        if Image.objects.filter(user=request.user).exists():
            image = Image.objects.get(user=request.user)
            image.image.delete()
            image.image = data
            image.save()
        else:
            image = Image(user=request.user, image=data)
            image.save()
        return JsonResponse({'status': 'ok'})


def get_user_name(request):
    return {'name': request.user.get_full_name(),
            'id': request.user.id}


@login_required(login_url='/user/login/')
def download_excel(request):
    def personal_excel(user):
        goals = list(Goal.objects.filter(owner_id=user,
                                         quarter=quarter,
                                         current=True).values())
        goals_count = len(goals)
        ws = wb.active
        ws.title = (user.get_full_name() or 'Sheet')
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        if request.user.is_superuser or request.user.has_perm('browse.change_goal'):
            ws['A1'] = '=HYPERLINK("#\'{}\'!A1", "{}")'.format('Общая сводка по коэф.', 'Главная')
            ws['A1'].style = 'Hyperlink'

        ws['A2'] = f'Сводка за {quarter} {user.get_full_name()}'
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws['A2'].border = thin_border
        ws.merge_cells('A2:H2')

        ws['A3'] = 'Название задачи'
        ws['A3'].alignment = Alignment(horizontal="center")
        ws['A3'].border = thin_border
        ws.merge_cells('A3:E3')

        ws['F3'] = 'Вес, %'
        ws['F3'].alignment = Alignment(horizontal="center")
        ws['F3'].border = thin_border

        ws['G3'] = 'Оценка, %'
        ws['G3'].alignment = Alignment(horizontal="center")
        ws.column_dimensions['G'].width = 13
        ws['G3'].border = thin_border

        ws['H3'] = 'Итог'
        ws['H3'].alignment = Alignment(horizontal="center")
        ws['H3'].border = thin_border

        ws['J3'] = 'Итоговый коэффициент:'
        ws['J3'].fill = PatternFill(start_color='FFC000',
                                    end_color='FFC000',
                                    fill_type='solid')
        ws['J3'].alignment = Alignment(horizontal="center")
        ws['J3'].border = Border(left=Side(style='medium'),
                                 top=Side(style='medium'),
                                 bottom=Side(style='medium'))
        ws.merge_cells('J3:L3')

        ws['M3'] = f'=SUM(H4:H{4+goals_count})'
        ws['M3'].alignment = Alignment(horizontal="right")
        ws['M3'].fill = PatternFill(start_color='D9D9D9',
                                    end_color='D9D9D9',
                                    fill_type='solid')
        ws['M3'].border = Border(right=Side(style='medium'),
                                 top=Side(style='medium'),
                                 bottom=Side(style='medium'))
        ws['M3'].number_format = '0.00'

        for row, goal in zip(ws[f'A4:A{goals_count+4}'], goals):
            for cell in row:
                cell.border = thin_border
                cell.value = goal['name']
                cell.alignment = Alignment(horizontal="center")
                number = int(''.join(filter(str.isdigit, cell.coordinate)))
                ws.merge_cells(f'{cell.coordinate}:E{number}')

        for row_idx, goal in enumerate(goals, start=4):
            cell = ws.cell(row=row_idx, column=6, value=goal['weight'])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
            cell = ws.cell(row=row_idx, column=7, value=goal['fact_mark'])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
            cell = ws.cell(row=row_idx, column=8,
                           value=f"=F{row_idx}/100 * G{row_idx}/100")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")

    wb = openpyxl.Workbook()
    quarter = request.GET.get('quarter')
    if request.user.is_superuser or request.user.has_perm('browse.change_goal'):
        ws = wb.active
        ws.title = 'Общая сводка по коэф.'
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ws['A2'] = f'Сводка за {quarter}'
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].border = thin_border
        ws.merge_cells('A2:H3')

        ws['A4'] = 'ФИО'
        ws['A4'].font = Font(bold=True, size=14)
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A4'].border = thin_border
        ws.merge_cells('A4:F5')

        ws['G4'] = 'Коэффициент'
        ws['G4'].font = Font(bold=True, size=14)
        ws['G4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G4'].border = thin_border
        ws.merge_cells('G4:H5')

        users = User.objects.all() if request.user.is_superuser else \
            User.objects.filter(groups__in=request.user.groups.all())
        i = 6
        for user in users:
            ws[f'A{i}'] = '=HYPERLINK("#\'{}\'!A1", "{}")'.format(user.get_full_name(), user.get_full_name())
            ws[f'A{i}'].style = 'Hyperlink'
            #link = f"workbookEx.xlsx#{user.get_full_name()}!E5"
            #ws[f'A{i}'].hyperlink = (link)
            ws[f'A{i}'].font = Font(bold=False, size=11)
            ws[f'A{i}'].alignment = Alignment(horizontal='center',
                                              vertical='center')
            ws[f'A{i}'].border = thin_border
            ws.merge_cells(f'A{i}:F{i+1}')
            goals = Goal.objects.filter(owner_id=user, quarter=quarter)
            koef = 0
            for goal in goals:
                koef += (goal.weight / 100) * (goal.fact_mark / 100)

            ws[f'G{i}'].number_format = '0.00'
            ws[f'G{i}'] = koef
            ws[f'G{i}'].font = Font(bold=True, size=11)
            ws[f'G{i}'].alignment = Alignment(horizontal='center',
                                              vertical='center')
            ws[f'G{i}'].border = thin_border
            ws.merge_cells(f'G{i}:H{i+1}')

            i += 2
        for user in users:
            ws = wb.create_sheet(user.get_full_name())
            wb.active = ws
            personal_excel(user)
    else:
        personal_excel(request.user)
    ws = wb.worksheets[0]
    wb.active = ws

    with NamedTemporaryFile(delete=True) as tmp_file:
        wb.save(tmp_file.name)
        title = f'Сводка за {quarter} {request.user.get_full_name()}'
        with open(tmp_file.name, 'rb') as f:
            response = HttpResponse(f.read(),
                                    content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = \
                f'attachment; filename*=UTF-8\'\'{quote(title)}.xlsx'
            return response
